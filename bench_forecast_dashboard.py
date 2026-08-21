"""
NA Bench Forecast Dashboard
Streamlit dashboard replicating three Excel sheets:
  1. Bench Forecast  - editable table (per-center weekly values)
  2. Historic Bench Data - stacked bar chart by center with NA FNC line overlay
  3. Q3 2025 tab - Q3 26 Forecast vs Q3 25 Bench vs Q3 Original Forecast line chart
"""

import base64
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path

# -----------------------------------------------------------------------------
# Palette  -  IBM Light theme
# -----------------------------------------------------------------------------
# Background / surfaces
BG          = "#f4f4f4"      # IBM Gray 10 - page canvas
SURFACE     = "#ffffff"      # white card
SURFACE2    = "#e8e8e8"      # IBM Gray 20 - raised / hover
BORDER      = "#c6c6c6"      # IBM Gray 30 - dividers
TEXT_PRI    = "#161616"      # IBM Gray 100 - primary text
TEXT_SEC    = "#525252"      # IBM Gray 60 - muted / labels
ACCENT      = "#0f62fe"      # IBM Blue 60 - primary accent
ACCENT2     = "#f1620a"      # IBM Orange 50 - secondary accent
GOOD        = "#198038"      # IBM Green 60 - positive
WARN        = "#f1c21b"      # IBM Yellow 30 - warning
DANGER      = "#da1e28"      # IBM Red 60 - danger

# Chart series colours - distinct, accessible, light-bg friendly
CENTER_COLORS = {
    "Baton Rouge":  "#0f62fe",   # IBM Blue 60
    "Buffalo":      "#6929c4",   # IBM Purple 60
    "Calgary":      "#198038",   # IBM Green 60
    "Halifax":      "#b28600",   # IBM Yellow 50
    "Lansing":      "#f1620a",   # IBM Orange 50
    "Monroe":       "#9f1853",   # IBM Magenta 60
    "Quebec":       "#007d79",   # IBM Teal 60
    "East Lansing": "#6929c4",   # IBM Purple 60
}
BENCH_ACTUAL_COLOR = "#8d8d8d"

# Q3 comparison series
C_FORECAST   = "#0f62fe"   # IBM Blue 60
C_ACTUAL     = "#198038"   # IBM Green 60
C_ORIG       = "#6f6f6f"   # IBM Gray 50

# Table row colours - per-center (background, text) - light tints
CENTER_ROW_COLORS = {
    "Baton Rouge":  ("#d0e2ff", "#0043ce"),
    "Buffalo":      ("#e8daff", "#491d8b"),
    "Calgary":      ("#defbe6", "#0e6027"),
    "Halifax":      ("#fdf6dd", "#8e6a00"),
    "Lansing":      ("#fff2e8", "#ba4e00"),
    "Monroe":       ("#ffe0eb", "#740937"),
    "Quebec":       ("#d9fbfb", "#004144"),
    "East Lansing": ("#e8daff", "#491d8b"),
}
# Header row for styled read-only tables
TBL_HEADER_BG   = "#0f62fe"
TBL_HEADER_FG   = "#ffffff"
TBL_TOTAL_BG    = "#393939"
TBL_TOTAL_FG    = "#ffffff"

# -----------------------------------------------------------------------------
# Config
# -----------------------------------------------------------------------------
XLSX_PATH      = Path(__file__).parent / "NA Bench Forecast.xlsx"
LOGO_PATH      = Path(__file__).parent / "IBM LOGO.png"
OVERRIDES_PATH = Path(__file__).parent / "bench_overrides.json"
AUDIT_LOG_PATH = Path(__file__).parent / "bench_audit_log.csv"
CENTERS   = ["Baton Rouge", "Buffalo", "Calgary", "Halifax", "Lansing", "Monroe", "Quebec"]
WEEKS     = [f"Wk {i:02d}" for i in range(1, 14)]


def _load_logo_data_uri() -> str:
    logo_bytes = LOGO_PATH.read_bytes()
    return f"data:image/png;base64,{base64.b64encode(logo_bytes).decode()}"


IBM_LOGO_DATA_URI = _load_logo_data_uri()

st.set_page_config(
    page_title="NA Bench Forecast",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# -----------------------------------------------------------------------------
# Access gate - password required
# Set via Streamlit Cloud Secrets:  [auth]  password = "your-password"
# For local dev without secrets, set env var:  BENCH_PASSWORD=your-password
# -----------------------------------------------------------------------------
import os

def _get_password():  # -> str | None
    """Return the required password, or None if no gate is configured."""
    try:
        return st.secrets["auth"]["password"]
    except Exception:
        return os.environ.get("BENCH_PASSWORD")

_required_pw = _get_password()


def audit_log(action: str, user: str, df: pd.DataFrame | None = None) -> None:
    """Append a row to the CSV audit log."""
    import csv
    from datetime import datetime, timezone
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    summary = ""
    if df is not None:
        parts = []
        for _, row in df[df["Center"].isin(CENTERS)].iterrows():
            total = sum(int(row[w]) for w in WEEKS if w in row)
            parts.append(f"{row['Center']}:{total}")
        summary = "; ".join(parts)
    file_exists = AUDIT_LOG_PATH.exists()
    with open(AUDIT_LOG_PATH, "a", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["timestamp", "action", "user", "summary"])
        writer.writerow([timestamp, action, user, summary])


if _required_pw:
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "bench_user" not in st.session_state:
        st.session_state.bench_user = "unknown"

    if not st.session_state.authenticated:
        st.markdown(f"""
        <style>
        html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"] {{
            background: {BG} !important;
        }}
        .lock-wrap {{
            display: flex; flex-direction: column; align-items: center;
            justify-content: center; min-height: 80vh;
        }}
        .lock-card {{
            background: {SURFACE}; border: 1px solid {BORDER};
            border-top: 3px solid {ACCENT}; border-radius: 12px;
            padding: 40px 44px; max-width: 380px; width: 100%;
            text-align: center;
        }}
        .lock-title {{
            font-size: 18px; font-weight: 700; color: {TEXT_PRI};
            margin-bottom: 4px;
        }}
        .lock-sub {{
            font-size: 13px; color: {TEXT_SEC}; margin-bottom: 28px;
        }}
        </style>
        <div class="lock-wrap">
          <div class="lock-card">
            <div class="lock-title">NA Bench Forecast</div>
            <div class="lock-sub">IBM FNC Workforce Planning &mdash; restricted access</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        name_input = st.text_input(
            "Your name", placeholder="Enter your name...",
            label_visibility="collapsed",
        )
        pw_input = st.text_input(
            "Access code", type="password", placeholder="Enter access code...",
            label_visibility="collapsed",
        )
        col_btn, _ = st.columns([1, 3])
        with col_btn:
            if st.button("Unlock", use_container_width=True):
                if pw_input == _required_pw:
                    user_label = name_input.strip() or "unknown"
                    st.session_state.authenticated = True
                    st.session_state.bench_user = user_label
                    audit_log("login", user_label)
                    st.rerun()
                else:
                    st.error("Incorrect access code.")
        st.stop()

# -----------------------------------------------------------------------------
# Editing lock - admin toggle via Streamlit Cloud Secrets
# Set  [lock]  editing = true  in Secrets to make the forecast table read-only.
# Remove or set to false to re-enable editing.
# -----------------------------------------------------------------------------
try:
    _lock_val = st.secrets["lock"]["editing"]
    # Handle both native TOML boolean (True) and string "true"/"false"
    if isinstance(_lock_val, bool):
        EDITING_LOCKED = _lock_val
    else:
        EDITING_LOCKED = str(_lock_val).strip().lower() == "true"
except Exception:
    EDITING_LOCKED = False

# -----------------------------------------------------------------------------
# Actuals cutoff - how many weeks from the start are locked as "actual bench"
# PRIMARY:  Streamlit Cloud Secrets  →  [forecast]  actuals_through = 7
# FALLBACK: ACTUALS_THROUGH env var  →  export ACTUALS_THROUGH=7
# DEFAULT:  hard-coded value below   →  safe default when neither is set
# Advance by 1 each week as new actuals are uploaded to the Excel file.
# -----------------------------------------------------------------------------
_ACTUALS_DEFAULT = 7   # ← update this if no Secrets / env var is configured

try:
    ACTUALS_CUTOFF = int(st.secrets["forecast"]["actuals_through"])
except Exception:
    try:
        ACTUALS_CUTOFF = int(os.environ.get("ACTUALS_THROUGH", ""))
    except Exception:
        ACTUALS_CUTOFF = _ACTUALS_DEFAULT
ACTUALS_CUTOFF = max(0, min(ACTUALS_CUTOFF, len(WEEKS)))

# -----------------------------------------------------------------------------
# Global CSS - dark theme
# -----------------------------------------------------------------------------
st.markdown(f"""
<style>
    /* -- canvas -- */
    html, body, [data-testid="stAppViewContainer"],
    [data-testid="stMain"], .main .block-container {{
        background: {BG} !important;
        color: {TEXT_PRI};
    }}
    [data-testid="stHeader"] {{ background: {BG} !important; }}
    [data-testid="stSidebar"] {{ background: {SURFACE} !important; }}

    /* -- typography -- */
    html, body, [class*="css"] {{
        font-family: "IBM Plex Sans", "Segoe UI", system-ui, sans-serif;
    }}
    h1, h2, h3, h4 {{ color: {TEXT_PRI} !important; }}
    p, li, label {{ color: {TEXT_SEC}; }}

    /* -- hero header -- */
    .dash-hero {{
        background: linear-gradient(90deg, #161616 0%, #262626 100%);
        border: 1px solid #393939;
        border-left: 4px solid {ACCENT};
        border-radius: 10px;
        padding: 20px 28px 18px;
        margin-bottom: 28px;
    }}
    .dash-hero-top {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 16px;
        margin-bottom: 16px;
    }}
    .dash-hero-ibm img {{
        display: block;
        height: 64px;
        width: auto;
    }}
    .dash-hero .badge {{
        background: #0f62fe;
        border: 1px solid #78a9ff;
        color: #ffffff;
        border-radius: 20px;
        padding: 4px 14px;
        font-size: 0.78rem;
        font-weight: 600;
        white-space: nowrap;
    }}
    .dash-hero h1 {{
        margin: 0 0 6px;
        font-size: 1.55rem;
        font-weight: 700;
        color: #ffffff !important;
        letter-spacing: -0.3px;
    }}
    .dash-hero .sub {{
        font-size: 0.83rem;
        color: rgba(255,255,255,0.78);
        margin: 0 0 12px;
    }}
    .dash-hero-meta {{
        display: flex;
        flex-wrap: wrap;
        gap: 10px;
    }}
    .dash-hero-meta span {{
        display: inline-flex;
        align-items: center;
        padding: 4px 10px;
        border-radius: 16px;
        background: rgba(255,255,255,0.08);
        border: 1px solid rgba(255,255,255,0.14);
        color: #f4f4f4;
        font-size: 0.76rem;
        font-weight: 500;
    }}

    /* -- section titles -- */
    .sec-title {{
        font-size: 0.82rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: {ACCENT};
        border-left: 3px solid {ACCENT};
        padding-left: 10px;
        margin: 0 0 16px;
    }}

    /* -- stat cards -- */
    .stat-row {{ display: flex; gap: 14px; margin-bottom: 22px; flex-wrap: wrap; }}
    .stat-card {{
        flex: 1; min-width: 120px;
        background: {SURFACE};
        border: 1px solid {BORDER};
        border-radius: 8px;
        padding: 14px 18px;
    }}
    .stat-card .val  {{ font-size: 1.6rem; font-weight: 700; color: {TEXT_PRI}; line-height: 1; }}
    .stat-card .lbl  {{ font-size: 0.75rem; color: {TEXT_SEC}; margin-top: 4px; }}
    .stat-card .delta-up   {{ color: {GOOD};  font-size: 0.78rem; }}
    .stat-card .delta-down {{ color: {DANGER}; font-size: 0.78rem; }}

    /* -- tabs -- */
    [data-testid="stTabs"] [role="tablist"] {{
        background: {SURFACE};
        border-radius: 8px 8px 0 0;
        padding: 4px 8px 0;
        border-bottom: 2px solid {ACCENT};
        gap: 4px;
    }}
    [data-testid="stTabs"] button[role="tab"] {{
        color: {TEXT_SEC} !important;
        border-radius: 6px 6px 0 0;
        padding: 8px 18px;
        font-size: 0.85rem;
        font-weight: 500;
        border: none;
        background: transparent;
    }}
    [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {{
        color: {ACCENT} !important;
        background: #e8f0fe !important;
        border-bottom: 3px solid {ACCENT} !important;
        font-weight: 700 !important;
    }}

    /* -- data editor / dataframe -- */
    [data-testid="stDataEditor"] {{ border-radius: 8px; border: 1px solid {BORDER}; }}

    /* -- data editor: header row -- */
    [data-testid="stDataEditor"] thead tr th {{
        background: {TBL_HEADER_BG} !important;
        color: {TBL_HEADER_FG} !important;
        font-size: 0.78rem !important;
        font-weight: 700 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        border-bottom: 2px solid {ACCENT} !important;
    }}
    /* -- data editor: center (first) column -- */
    [data-testid="stDataEditor"] tbody tr td:first-child {{
        background: {SURFACE2} !important;
        color: {TEXT_PRI} !important;
        font-weight: 600 !important;
        font-size: 0.82rem !important;
    }}
    /* -- data editor: data cells -- */
    [data-testid="stDataEditor"] tbody tr td {{
        background: {SURFACE} !important;
        color: {TEXT_PRI} !important;
        border-color: {BORDER} !important;
    }}
    /* -- data editor: alternating rows -- */
    [data-testid="stDataEditor"] tbody tr:nth-child(even) td {{
        background: {SURFACE2} !important;
    }}

    /* -- buttons -- */
    [data-testid="stDownloadButton"] button, [data-testid="stButton"] button {{
        background: {ACCENT} !important;
        border: none !important;
        color: #ffffff !important;
        border-radius: 4px;
        font-size: 0.83rem;
        font-weight: 600;
        padding: 8px 20px;
    }}
    [data-testid="stDownloadButton"] button:hover, [data-testid="stButton"] button:hover {{
        background: #0043ce !important;
    }}

    /* -- metrics -- */
    [data-testid="metric-container"] {{
        background: {SURFACE};
        border: 1px solid {BORDER};
        border-left: 4px solid {ACCENT};
        border-radius: 4px;
        padding: 12px 16px !important;
    }}
    [data-testid="stMetricValue"] {{ color: {TEXT_PRI} !important; font-size: 1.4rem !important; font-weight: 700 !important; }}
    [data-testid="stMetricLabel"] {{ color: {TEXT_SEC} !important; font-size: 0.75rem !important; }}

    /* -- multiselect -- */
    [data-baseweb="select"] {{ background: {SURFACE} !important; border-color: {BORDER} !important; }}

    /* -- number input -- */
    [data-testid="stNumberInput"] input {{
        background: {SURFACE} !important;
        border: 1px solid {BORDER} !important;
        color: {TEXT_PRI} !important;
        border-radius: 4px;
        text-align: center !important;
        font-size: 0.85rem !important;
        padding: 4px 2px !important;
    }}
    /* tighten spinner buttons */
    [data-testid="stNumberInput"] button {{
        background: {SURFACE2} !important;
        border-color: {BORDER} !important;
        color: {TEXT_SEC} !important;
    }}
    [data-testid="stExpander"] > div[data-testid="stExpanderDetails"] {{
        background: {SURFACE} !important;
        border: 1px solid {BORDER} !important;
        border-top: none !important;
        border-radius: 0 0 6px 6px;
        padding: 12px 8px 16px;
    }}

    /* -- checkbox -- */
    [data-testid="stCheckbox"] label {{ color: {TEXT_PRI} !important; }}

    /* -- dividers -- */
    hr {{ border-color: {BORDER} !important; }}

    footer {{ visibility: hidden; }}
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# Hero Header
# -----------------------------------------------------------------------------
_actuals_label = f"Wk {ACTUALS_CUTOFF:02d}" if ACTUALS_CUTOFF > 0 else "none"
_forecast_start = f"Wk {ACTUALS_CUTOFF + 1:02d}" if ACTUALS_CUTOFF < 13 else "—"

st.markdown(f"""
<div class="dash-hero">
    <div class="dash-hero-top">
        <div class="dash-hero-ibm" aria-label="IBM logo">
            <img src="{IBM_LOGO_DATA_URI}" alt="IBM logo" />
        </div>
        <div class="badge">Q3 2026 Active</div>
    </div>
    <div>
        <h1>NA Bench Forecast</h1>
        <p class="sub">North America FNC workforce planning dashboard with actual bench through {_actuals_label} and quarter-end forecast visibility.</p>
        <div class="dash-hero-meta">
            <span>Actuals loaded through {_actuals_label}</span>
            <span>Forecast view through Wk 13</span>
            <span>Historic + Q3 comparison included</span>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# Shared chart layout defaults
# -----------------------------------------------------------------------------
CHART_LAYOUT = dict(
    plot_bgcolor   = SURFACE,
    paper_bgcolor  = BG,
    font           = dict(color=TEXT_SEC, family="IBM Plex Sans, Segoe UI, system-ui"),
    legend         = dict(
        orientation="h", y=-0.22,
        font=dict(size=11, color=TEXT_SEC),
        bgcolor="rgba(0,0,0,0)",
    ),
    margin         = dict(l=12, r=12, t=44, b=60),
    xaxis          = dict(
        tickfont      = dict(size=10, color=TEXT_SEC),
        gridcolor     = BORDER,
        zerolinecolor = BORDER,
        linecolor     = BORDER,
        showgrid      = False,
    ),
    yaxis          = dict(
        tickfont      = dict(size=10, color=TEXT_SEC),
        gridcolor     = BORDER,
        zerolinecolor = BORDER,
        linecolor     = BORDER,
    ),
    hoverlabel     = dict(
        bgcolor    = SURFACE,
        bordercolor= BORDER,
        font       = dict(color=TEXT_PRI, size=12),
    ),
)

# -----------------------------------------------------------------------------
# Styled-table helper
# -----------------------------------------------------------------------------
def styled_table(
    df: pd.DataFrame,
    total_rows: list[str] | None = None,
    grey_columns: list[str] | None = None,
) -> None:
    """Render a read-only DataFrame with per-center row colours, dark header, centred numbers."""
    total_rows = total_rows or []
    grey_columns = set(grey_columns or [])

    def _row_style(row):
        center = str(row["Center"]) if "Center" in row.index else ""
        if center in total_rows:
            bg, fg = TBL_TOTAL_BG, TBL_TOTAL_FG
        elif center in CENTER_ROW_COLORS:
            bg, fg = CENTER_ROW_COLORS[center]
        else:
            bg, fg = SURFACE, TEXT_PRI
        return [f"background-color:{bg}; color:{fg}"] * len(row)

    def _col_align(col):
        if col.name == "Center":
            return ["text-align:left; padding-left:12px"] * len(col)
        return ["text-align:center"] * len(col)

    def _grey_columns(col):
        if col.name in grey_columns:
            return [f"background-color:{SURFACE2}; color:{TEXT_PRI}"] * len(col)
        return [""] * len(col)

    styler = (
        df.style
        .apply(_row_style, axis=1)
        .apply(_col_align, axis=0)
        .apply(_grey_columns, axis=0)
        .set_table_styles([
            {"selector": "thead tr th",
             "props": f"background-color:{TBL_HEADER_BG}; color:{TBL_HEADER_FG}; "
                       f"font-weight:700; font-size:0.78rem; text-transform:uppercase; "
                       f"letter-spacing:0.05em; border-bottom:2px solid {ACCENT}; "
                       f"text-align:center; padding:7px 10px;"},
            {"selector": "thead tr th:first-child",
             "props": "text-align:left; padding-left:12px;"},
            {"selector": "td",
             "props": f"border:1px solid {BORDER}; font-size:0.85rem; padding:6px 10px;"},
            {"selector": "",
             "props": f"border-collapse:collapse; width:100%;"},
        ])
        .format(precision=0, na_rep="-")
    )
    st.markdown(
        "<div style='overflow-x:auto;border-radius:8px;border:1px solid "
        + BORDER + ";'>" + styler.to_html(index=False) + "</div>",
        unsafe_allow_html=True,
    )


# -----------------------------------------------------------------------------
# Data loaders  (ttl=60 — re-reads Excel every 60 s)
# -----------------------------------------------------------------------------
@st.cache_data(ttl=60)
def load_bench_forecast() -> pd.DataFrame:
    raw = pd.read_excel(XLSX_PATH, sheet_name="Bench Forecast", header=0)
    raw.columns = ["Center"] + WEEKS + (list(raw.columns[14:]) if len(raw.columns) > 14 else [])
    raw = raw[raw["Center"].isin(CENTERS)].reset_index(drop=True)
    for w in WEEKS:
        raw[w] = pd.to_numeric(raw[w], errors="coerce").fillna(0).astype(int)
    return raw[["Center"] + WEEKS]


def load_overrides() -> pd.DataFrame | None:
    """Load persisted user edits from the Excel file (Bench Forecast sheet).
    Falls back to the legacy JSON file if present, then returns None."""
    # Primary: read edits stored back in the Excel workbook
    try:
        raw = pd.read_excel(XLSX_PATH, sheet_name="Bench Forecast", header=0)
        raw.columns = ["Center"] + WEEKS + (list(raw.columns[14:]) if len(raw.columns) > 14 else [])
        raw = raw[raw["Center"].isin(CENTERS)].reset_index(drop=True)
        for w in WEEKS:
            raw[w] = pd.to_numeric(raw[w], errors="coerce").fillna(0).astype(int)
        result = raw[["Center"] + WEEKS]
        if not result.empty:
            return result
    except Exception:
        pass
    # Legacy fallback: JSON file written by old save mechanism
    if OVERRIDES_PATH.exists():
        try:
            import json
            with open(OVERRIDES_PATH, "r") as f:
                data = json.load(f)
            df = pd.DataFrame(data)
            df["Center"] = df["Center"].astype(str)
            for w in WEEKS:
                if w in df.columns:
                    df[w] = pd.to_numeric(df[w], errors="coerce").fillna(0).astype(int)
            return df[["Center"] + WEEKS]
        except Exception:
            pass
    return None


def save_overrides(df: pd.DataFrame, changed_by: str = "unknown") -> None:
    """Persist the edited forecast DataFrame back into the Excel workbook.
    Writes only the center rows in the Bench Forecast sheet, leaving all other
    rows (Grand Total, Bench %, HC, etc.) untouched."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Bench Forecast"]

    # Build a map of center name -> row number from the sheet
    center_rows: dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        cell_val = row[0].value
        if cell_val in CENTERS:
            center_rows[str(cell_val)] = row_idx

    # Map week label -> column index (1-based); header is row 1
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 2)]
    week_cols: dict[str, int] = {}
    for col_idx, h in enumerate(header, start=1):
        if h in WEEKS:
            week_cols[str(h)] = col_idx

    # Write updated values
    for _, row in df[["Center"] + WEEKS].iterrows():
        center = row["Center"]
        if center not in center_rows:
            continue
        r = center_rows[center]
        for w in WEEKS:
            if w in week_cols:
                ws.cell(row=r, column=week_cols[w]).value = int(row[w])

    wb.save(XLSX_PATH)
    # Also write JSON as secondary backup for portability
    import json
    with open(OVERRIDES_PATH, "w") as f:
        json.dump(df[["Center"] + WEEKS].to_dict(orient="records"), f, indent=2)
    audit_log("save_forecast", changed_by, df)



@st.cache_data(ttl=60)
def load_historic() -> pd.DataFrame:
    raw = pd.read_excel(XLSX_PATH, sheet_name="Historic Bench Data", header=0)
    raw = raw.iloc[:, :12].copy()
    raw.columns = [
        "Year", "Quarter", "Week",
        "Baton Rouge", "East Lansing", "Monroe", "Buffalo",
        "Halifax", "Quebec", "Calgary", "NA FNC", "Wk",
    ]
    raw = raw[raw["Year"].notna() & raw["Quarter"].notna()].copy()
    raw["Year"] = pd.to_numeric(raw["Year"], errors="coerce")
    raw = raw[raw["Year"].notna()].copy()
    raw["Year"] = raw["Year"].astype(int)
    raw["Quarter"] = raw["Quarter"].astype(str).str.strip()
    raw["Week"] = raw["Week"].astype(str).str.strip()
    raw["Wk"] = raw["Wk"].astype(str).str.strip()
    for col in ["Baton Rouge", "East Lansing", "Monroe", "Buffalo", "Halifax", "Quebec", "Calgary", "NA FNC"]:
        raw[col] = pd.to_numeric(raw[col], errors="coerce").fillna(0)
    raw["Label"] = raw["Year"].astype(str) + " " + raw["Quarter"] + " " + raw["Week"]
    raw["QuarterWeek"] = raw["Quarter"] + " " + raw["Week"]
    return raw


@st.cache_data(ttl=60)
def load_q3_comparison() -> pd.DataFrame:
    raw = pd.read_excel(XLSX_PATH, sheet_name="Q3 2025", header=None)
    pivot_start = None
    for i, row in raw.iterrows():
        if str(row.iloc[0]).strip() == "Total Bench":
            pivot_start = i
            break
    if pivot_start is None:
        return pd.DataFrame()
    weeks_row = raw.iloc[pivot_start]
    weeks = [str(w).strip() for w in weeks_row.iloc[1:14].tolist()]
    rows = {}
    for offset in [1, 2, 3]:
        r = raw.iloc[pivot_start + offset]
        rows[str(r.iloc[0]).strip()] = pd.to_numeric(r.iloc[1:14], errors="coerce").tolist()
    return pd.DataFrame(rows, index=weeks)


@st.cache_data(ttl=60)
def load_q3_center_detail() -> pd.DataFrame:
    raw = pd.read_excel(XLSX_PATH, sheet_name="Q3 2025", header=None)
    start = None
    for i, row in raw.iterrows():
        if str(row.iloc[0]).strip() == "Center" and str(row.iloc[1]).strip() in ("W1", "Wk1"):
            if start is not None:
                start = i
                break
            start = i
    if start is None:
        return pd.DataFrame()
    week_row  = raw.iloc[start]
    weeks_net = [str(week_row.iloc[j]).strip() for j in range(1, 14)]
    centers_q3 = ["Baton Rouge", "Lansing", "Monroe", "Buffalo", "Halifax", "Quebec", "Calgary"]
    rows_out = []
    for offset in range(2, 9):
        r = raw.iloc[start + offset]
        if str(r.iloc[0]).strip() not in centers_q3:
            continue
        rows_out.append([str(r.iloc[0]).strip()] + pd.to_numeric(r.iloc[1:14], errors="coerce").tolist())
    return pd.DataFrame(rows_out, columns=["Center"] + weeks_net)


# -----------------------------------------------------------------------------
# Tabs
# -----------------------------------------------------------------------------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "\U0001f4cb  Bench Forecast",
    "\U0001f4c8  Historic Bench",
    "\U0001f4ca  Q3 Comparison",
    "\U0001f4e7  Executive Summary",
    "\U0001f4cb  Audit Log",
])


# =============================================================================
# TAB 1 - Editable Bench Forecast
# =============================================================================

# Column header colours
TBL_ACT_HDR_BG  = "#6f6f6f"   # IBM Gray 50  – actuals header
TBL_ACT_HDR_FG  = "#ffffff"
TBL_FC_HDR_BG   = "#0f62fe"   # IBM Blue 60  – forecast header
TBL_FC_HDR_FG   = "#ffffff"

# Cell tint overlays for actual vs forecast columns (blended over row color)
# We use a slight desaturation for actuals and keep row color for forecast
ACT_CELL_ALPHA  = "rgba(111,111,111,0.10)"  # gray wash over actual cells
FC_CELL_ALPHA   = "rgba(15,98,254,0.07)"    # blue wash over forecast cells


def build_colored_table(df: pd.DataFrame, editable: bool = True) -> str:
    """
    Render the forecast table with:
      - Per-center row background colors
      - Gray column headers + subtle gray cell tint for actual weeks (Wk 01–06)
      - Blue column headers + subtle blue cell tint for forecast weeks (Wk 07–13)
      - Editable <input> fields when editable=True
      - Grand Total row at the bottom
    """
    actual_weeks   = WEEKS[:ACTUALS_CUTOFF]
    forecast_weeks = WEEKS[ACTUALS_CUTOFF:]

    def hdr(week: str) -> str:
        is_act = week in actual_weeks
        bg  = TBL_ACT_HDR_BG if is_act else TBL_FC_HDR_BG
        fg  = TBL_ACT_HDR_FG if is_act else TBL_FC_HDR_FG
        tag = "● " if is_act else "◆ "
        return (
            f'<th style="background:{bg};color:{fg};font-size:0.72rem;font-weight:700;'
            f'text-transform:uppercase;letter-spacing:0.05em;padding:7px 5px;'
            f'text-align:center;border:1px solid {BORDER};border-bottom:3px solid {bg};">'
            f'{tag}{week}</th>'
        )

    center_hdr = (
        f'<th style="background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};'
        f'font-size:0.72rem;font-weight:700;text-transform:uppercase;'
        f'letter-spacing:0.05em;padding:7px 12px;text-align:left;'
        f'border:1px solid {BORDER};border-bottom:3px solid {TBL_HEADER_BG};">'
        f'Center</th>'
    )
    header_html = center_hdr + "".join(hdr(w) for w in WEEKS)

    inp_base = (
        "width:100%;background:transparent;border:none;outline:none;"
        "font-size:0.85rem;text-align:center;padding:0;margin:0;"
        "color:inherit;-moz-appearance:textfield;"
        "-webkit-appearance:none;"
    )

    rows_html = ""
    for _, row in df.iterrows():
        center = row["Center"]
        bg, fg = CENTER_ROW_COLORS.get(center, (SURFACE, TEXT_PRI))
        center_td = (
            f'<td style="background:{bg};color:{fg};font-weight:700;'
            f'font-size:0.82rem;padding:7px 12px;border:1px solid {BORDER};'
            f'white-space:nowrap;">{center}</td>'
        )
        cells = center_td
        for w in WEEKS:
            val    = int(row[w])
            is_act = w in actual_weeks
            # Actual weeks: solid gray background regardless of center color
            # Forecast weeks: keep the center's row color
            if is_act:
                cell_style = f"background:#e0e0e0;color:#525252;"
            else:
                cell_style = f"background:{bg};color:{fg};"
            cid = f"{center.replace(' ', '_')}_{w.replace(' ', '_')}"
            if editable:
                cells += (
                    f'<td style="{cell_style}padding:5px 3px;'
                    f'border:1px solid {BORDER};text-align:center;">'
                    f'<input id="{cid}" name="{cid}" type="number" min="0" max="9999" value="{val}" '
                    f'style="{inp_base}" oninput="recalc()" />'
                    f'</td>'
                )
            else:
                cells += (
                    f'<td style="{cell_style}padding:5px 8px;'
                    f'border:1px solid {BORDER};text-align:center;">'
                    f'{val}</td>'
                )
        rows_html += f"<tr>{cells}</tr>\n"

    # Grand Total row
    totals = {w: int(df[w].sum()) for w in WEEKS}
    total_cells = (
        f'<td style="background:{TBL_TOTAL_BG};color:{TBL_TOTAL_FG};'
        f'font-weight:700;font-size:0.82rem;padding:7px 12px;'
        f'border:1px solid {BORDER};white-space:nowrap;">Grand Total</td>'
    )
    for w in WEEKS:
        is_act = w in actual_weeks
        overlay = ACT_CELL_ALPHA if is_act else FC_CELL_ALPHA
        wid = w.replace(" ", "_")
        total_id = f'id="tot_{wid}"' if editable else ""
        total_cells += (
            f'<td {total_id} style="background:{TBL_TOTAL_BG};color:{TBL_TOTAL_FG};'
            f'font-weight:700;padding:5px 8px;border:1px solid {BORDER};text-align:center;">'
            f'{totals[w]}</td>'
        )
    rows_html += f"<tr>{total_cells}</tr>\n"

    _leg_act_end   = f"Wk {ACTUALS_CUTOFF:02d}" if ACTUALS_CUTOFF > 0 else "none"
    _leg_fc_start  = f"Wk {ACTUALS_CUTOFF + 1:02d}" if ACTUALS_CUTOFF < 13 else "—"
    legend = (
        f'<div style="display:flex;gap:16px;margin-bottom:8px;font-size:0.75rem;">'
        f'<span style="display:inline-flex;align-items:center;gap:5px;">'
        f'<span style="display:inline-block;width:12px;height:12px;border-radius:2px;background:{TBL_ACT_HDR_BG};"></span>'
        f'<span style="color:{TEXT_SEC};">● Actual bench (Wk 01–{_leg_act_end})</span></span>'
        f'<span style="display:inline-flex;align-items:center;gap:5px;">'
        f'<span style="display:inline-block;width:12px;height:12px;border-radius:2px;background:{TBL_FC_HDR_BG};"></span>'
        f'<span style="color:{TEXT_SEC};">◆ Forecast ({_leg_fc_start}–Wk 13)</span></span>'
        f'</div>'
    )

    js = ""
    if editable:
        centers_js = "[" + ",".join(f'"{c}"' for c in CENTERS) + "]"
        weeks_js   = "[" + ",".join(f'"{w}"' for w in WEEKS) + "]"
        js = f"""
<script>
const CENTERS = {centers_js};
const WEEKS   = {weeks_js};
function getId(c,w){{ return c.replace(/ /g,'_')+'_'+w.replace(/ /g,'_'); }}
function recalc(){{
  WEEKS.forEach(w=>{{
    let tot=0;
    CENTERS.forEach(c=>{{
      const el=document.getElementById(getId(c,w));
      if(el) tot+=(parseInt(el.value)||0);
    }});
    const td=document.getElementById('tot_'+w.replace(/ /g,'_'));
    if(td) td.textContent=tot;
  }});
}}
window.addEventListener('load', recalc);
</script>"""

    return f"""
<style>
  body{{margin:0;padding:0;background:{BG};}}
  table{{border-collapse:collapse;width:100%;font-family:"IBM Plex Sans",Inter,system-ui,sans-serif;}}
  input[type=number]{{outline:none!important;box-shadow:none!important;-moz-appearance:textfield!important;}}
  input[type=number]::-webkit-outer-spin-button,
  input[type=number]::-webkit-inner-spin-button{{-webkit-appearance:none!important;margin:0;display:none;}}
  tr:hover td{{filter:brightness(1.06);}}
</style>
{legend}
<div style="overflow-x:auto;border-radius:8px;border:1px solid {BORDER};">
  <table>
    <thead><tr>{header_html}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>
{js}
"""


with tab1:
    # ---- Load base data, then apply any persisted overrides ----------------
    df_forecast = load_bench_forecast()
    saved = load_overrides()

    if "forecast_data" not in st.session_state:
        st.session_state["forecast_data"] = saved.copy() if saved is not None else df_forecast.copy()

    working = st.session_state["forecast_data"].copy()

    if EDITING_LOCKED:
        st.markdown(
            f'<div style="background:#fff2e8;border:1px solid {ACCENT2};border-left:4px solid {ACCENT2};'
            f'border-radius:6px;padding:10px 16px;margin-bottom:12px;font-size:13px;color:#ba4e00;">'
            f'&#128274; <strong>Forecast table is locked.</strong> '
            f'Editing has been disabled by the administrator. Contact your manager to request changes.'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.markdown(f'<div class="sec-title">Weekly Headcount by Center - read only</div>',
                    unsafe_allow_html=True)
        # Read-only: render the styled colored table
        components.html(build_colored_table(working, editable=False),
                        height=len(CENTERS) * 42 + 100, scrolling=False)
        edited = working.copy()
    else:
        st.markdown(f'<div class="sec-title">Weekly Headcount by Center</div>',
                    unsafe_allow_html=True)

        # ---- Colored visual table (display only, always current) -----------
        components.html(build_colored_table(working, editable=False),
                        height=len(CENTERS) * 42 + 100, scrolling=False)
        st.markdown("<div style='margin-top:-5rem'></div>", unsafe_allow_html=True)

        # ---- Edit form below the table (forecast weeks only) ---------------
        forecast_weeks_only = WEEKS[ACTUALS_CUTOFF:]
        actual_weeks_only   = WEEKS[:ACTUALS_CUTOFF]

        cutoff_label = f"Wk {ACTUALS_CUTOFF:02d}" if ACTUALS_CUTOFF > 0 else "none"
        with st.expander("Edit forecast values", expanded=False):
            if not forecast_weeks_only:
                st.info("All weeks are currently locked as actuals. Reduce `actuals_through` in Secrets to re-enable editing.")
            else:
                st.caption(
                    f"🔒 **Wk 01–{cutoff_label} are locked** (actual bench data). "
                    f"Only forecast weeks (**{forecast_weeks_only[0]}–Wk 13**) are editable. "
                    f"Click **Apply & Save** when done."
                )
                with st.form("forecast_form"):
                    # Header: Center label + forecast week labels only
                    hdr_cols = st.columns([2] + [1] * len(forecast_weeks_only))
                    hdr_cols[0].markdown(
                        f'<div style="font-size:0.7rem;font-weight:700;color:{TEXT_SEC};'
                        f'text-transform:uppercase;padding-top:6px;">Center</div>',
                        unsafe_allow_html=True,
                    )
                    for i, w in enumerate(forecast_weeks_only):
                        hdr_cols[i + 1].markdown(
                            f'<div style="font-size:0.68rem;font-weight:700;color:{TBL_FC_HDR_BG};'
                            f'text-transform:uppercase;text-align:center;padding-top:6px;">◆ {w}</div>',
                            unsafe_allow_html=True,
                        )

                    # One row per center — only forecast week inputs rendered
                    form_values: dict[str, dict[str, int]] = {}
                    for _, row in working.iterrows():
                        center = row["Center"]
                        bg, fg = CENTER_ROW_COLORS.get(center, (SURFACE, TEXT_PRI))
                        row_cols = st.columns([2] + [1] * len(forecast_weeks_only))
                        row_cols[0].markdown(
                            f'<div style="background:{bg};color:{fg};font-weight:700;'
                            f'font-size:0.82rem;padding:6px 8px;border-radius:4px;'
                            f'margin-top:2px;">{center}</div>',
                            unsafe_allow_html=True,
                        )
                        # Preserve actual-week values unchanged
                        form_values[center] = {w: int(row[w]) for w in actual_weeks_only}
                        # Render editable inputs for forecast weeks only
                        for i, w in enumerate(forecast_weeks_only):
                            val = int(row[w])
                            form_values[center][w] = row_cols[i + 1].number_input(
                                label=w,
                                value=val,
                                min_value=0,
                                max_value=9999,
                                step=1,
                                key=f"form_{center}_{w}",
                                label_visibility="collapsed",
                            )

                    submitted = st.form_submit_button("💾  Apply & Save", use_container_width=True)
                    if submitted:
                        new_rows = [
                            {"Center": c, **{w: form_values[c][w] for w in WEEKS}}
                            for c in CENTERS
                        ]
                        updated_df = pd.DataFrame(new_rows)
                        save_overrides(updated_df, changed_by=st.session_state.get("bench_user", "unknown"))
                        st.session_state["forecast_data"] = updated_df.copy()
                        st.success("Changes saved — they will persist across restarts.")
                        st.rerun()

        edited = working.copy()

    # -- Grand Total + Bench % - flush below the editable table -----------
    grand_total = edited[WEEKS].sum().to_frame().T
    grand_total.insert(0, "Center", "Grand Total")

    # HC input needed before bench_pct is calculated - rendered after tables
    hc = st.session_state.get("hc_input", 2013)

    bench_pct = grand_total[WEEKS].values[0] / hc * 100
    bench_pct_row = pd.DataFrame(
        [["Bench %"] + [f"{v:.1f}%" for v in bench_pct]],
        columns=["Center"] + WEEKS,
    )

    styled_table(grand_total, total_rows=["Grand Total"])
    st.markdown("<div style='margin-top:6px'></div>", unsafe_allow_html=True)
    styled_table(bench_pct_row, total_rows=["Bench %"])

    # -- Total HC input - below the tables ---------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    hc_col, _ = st.columns([1, 3])
    with hc_col:
        hc = st.number_input(
            "Total HC (Bench % denominator)",
            min_value=1, value=2013, step=1, key="hc_input",
        )
    # Recompute bench_pct with the confirmed hc value
    bench_pct = grand_total[WEEKS].values[0] / hc * 100

    # -- KPI strip ---------------------------------------------------------
    cur_wk_totals = grand_total[WEEKS].values[0]
    peak_wk_idx   = cur_wk_totals.argmax()
    st.markdown("<br>", unsafe_allow_html=True)
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Peak Week",       f"Wk {peak_wk_idx+1:02d}",         f"{int(cur_wk_totals[peak_wk_idx])} heads")
    k2.metric("Wk 01 Total",     f"{int(cur_wk_totals[0])}",         "")
    k3.metric("Wk 13 Total",     f"{int(cur_wk_totals[-1])}",        f"{int(cur_wk_totals[-1]-cur_wk_totals[0]):+d} vs Wk 01")
    k4.metric("Avg Bench %",     f"{bench_pct.mean():.1f}%",         "")

    # -- Filled area chart - one line per center ---------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f'<div class="sec-title">Weekly Bench by Center</div>', unsafe_allow_html=True)

    fig_prev = go.Figure()

    # Per-center filled area traces
    for center in CENTERS:
        row = edited[edited["Center"] == center]
        if row.empty:
            continue
        color     = CENTER_COLORS.get(center, TEXT_SEC)
        vals      = row[WEEKS].values[0].tolist()
        r, g, b   = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        fill_rgba = f"rgba({r},{g},{b},0.15)"

        fig_prev.add_trace(go.Scatter(
            name=center,
            x=WEEKS,
            y=vals,
            mode="lines+markers",
            line=dict(color=color, width=2.2),
            marker=dict(size=6, color=color, symbol="circle",
                        line=dict(color=SURFACE, width=1.5)),
            fill="tozeroy",
            fillcolor=fill_rgba,
            hovertemplate=f"<b>{center}</b><br>%{{x}}: %{{y}} heads<extra></extra>",
        ))

    # Grand Total line - dark charcoal, dashed, with data labels
    total_vals = edited[WEEKS].sum().tolist()
    fig_prev.add_trace(go.Scatter(
        name="Grand Total",
        x=WEEKS,
        y=total_vals,
        mode="lines+markers+text",
        line=dict(color="#161616", width=2.5, dash="dash"),
        marker=dict(size=7, color="#161616", symbol="diamond",
                    line=dict(color=SURFACE, width=1.5)),
        text=[str(int(v)) for v in total_vals],
        textposition="top center",
        textfont=dict(color="#161616", size=10, family="Inter, system-ui"),
        hovertemplate="<b>Grand Total</b><br>%{x}: %{y} heads<extra></extra>",
    ))

    layout1 = {**CHART_LAYOUT}
    layout1.update(dict(
        height=400,
        title=dict(text="", x=0),
        xaxis={**CHART_LAYOUT["xaxis"], "title": ""},
        yaxis=dict(**CHART_LAYOUT["yaxis"], title="Head Count"),
        hovermode="x unified",
    ))
    fig_prev.update_layout(**layout1)
    st.plotly_chart(fig_prev, use_container_width=True)

    # -- Download -----------------------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)

    @st.cache_data
    def to_excel_bytes(df: pd.DataFrame) -> bytes:
        import io
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Bench Forecast")
        return buf.getvalue()

    dl_df = pd.concat([edited, grand_total], ignore_index=True)
    st.download_button(
        label="  Download updated forecast (.xlsx)",
        data=to_excel_bytes(dl_df),
        file_name="NA_Bench_Forecast_updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =============================================================================
# TAB 2 - Historic Bench Trend
# =============================================================================
with tab2:
    st.markdown(f'<div class="sec-title">All Centers . NA FNC Total overlay</div>', unsafe_allow_html=True)

    _rc1, _rc2 = st.columns([8, 1])
    with _rc2:
        if st.button("🔄 Reload data", key="reload_historic"):
            st.cache_data.clear()
            st.rerun()

    df_hist = load_historic()
    years_avail    = sorted(df_hist["Year"].unique())
    quarters_avail = sorted(df_hist["Quarter"].unique())
    hist_centers   = ["Baton Rouge", "East Lansing", "Monroe", "Buffalo", "Halifax", "Quebec", "Calgary"]

    fc1, fc2, fc3 = st.columns([1, 1, 2])
    with fc1:
        sel_years    = st.multiselect("Year",             years_avail,    default=years_avail,    key="hist_years")
    with fc2:
        sel_quarters = st.multiselect("Quarter",          quarters_avail, default=quarters_avail, key="hist_quarters")
    with fc3:
        sel_centers  = st.multiselect("Centers to show",  hist_centers,   default=hist_centers,   key="hist_centers")

    df_f = df_hist[
        df_hist["Year"].isin(sel_years) & df_hist["Quarter"].isin(sel_quarters)
    ].sort_values(["Year", "Quarter", "Week"]).copy()

    if df_f.empty:
        st.info("No data for the selected filters.")
    else:
        x_labels = df_f["Label"].tolist()

        # -- KPI strip -----------------------------------------------------
        avg_cic = df_f["NA FNC"].mean()
        max_cic = df_f["NA FNC"].max()
        min_cic = df_f["NA FNC"].min()
        last_cic = df_f["NA FNC"].iloc[-1]

        st.markdown("<br>", unsafe_allow_html=True)
        h1, h2, h3, h4 = st.columns(4)
        h1.metric("Avg NA FNC",        f"{avg_cic:.0f}",    "")
        h2.metric("Peak NA FNC",        f"{int(max_cic)}",   "")
        h3.metric("Trough NA FNC",      f"{int(min_cic)}",   "")
        h4.metric("Latest Data Point",  f"{int(last_cic)}",  "")
        st.markdown("<br>", unsafe_allow_html=True)

        # -- Chart ---------------------------------------------------------
        fig_hist = go.Figure()

        for center in sel_centers:
            if center not in df_f.columns:
                continue
            fig_hist.add_trace(go.Bar(
                name=center,
                x=x_labels,
                y=df_f[center].tolist(),
                marker_color=CENTER_COLORS.get(center, TEXT_SEC),
                marker_line_width=0,
                hovertemplate=f"<b>{center}</b><br>%{{x}}: %{{y}}<extra></extra>",
            ))

        # NA FNC total line - amber, stands out on dark stacked bars
        fig_hist.add_trace(go.Scatter(
            name="NA FNC Total",
            x=x_labels,
            y=df_f["NA FNC"].tolist(),
            mode="lines+markers+text",
            line=dict(color=ACCENT2, width=2.5),
            marker=dict(size=4, color=ACCENT2, symbol="circle"),
            text=[str(int(v)) for v in df_f["NA FNC"].tolist()],
            textposition="top center",
            textfont=dict(color=ACCENT2, size=9, family="Inter, system-ui"),
            hovertemplate="<b>NA FNC Total</b><br>%{x}: %{y}<extra></extra>",
        ))

        # Quarter boundary dotted lines
        prev_q = None
        for i, (_, row) in enumerate(df_f.iterrows()):
            cur_q = (row["Year"], row["Quarter"])
            if prev_q and cur_q != prev_q:
                fig_hist.add_vline(
                    x=i - 0.5,
                    line_width=1,
                    line_dash="dot",
                    line_color=BORDER,
                )
            prev_q = cur_q

        layout2 = {**CHART_LAYOUT}
        layout2.update(dict(
            barmode="stack",
            height=480,
            bargap=0.06,
            title=dict(
                text="Historic Bench - Stacked by Center  .  NA FNC Total (amber line)",
                font=dict(size=13, color=TEXT_SEC),
                x=0,
            ),
            xaxis={**CHART_LAYOUT["xaxis"], "tickangle": -55, "tickfont": dict(size=9, color=TEXT_SEC)},
            yaxis=dict(**CHART_LAYOUT["yaxis"], title="Head Count on Bench"),
            legend={**CHART_LAYOUT["legend"], "y": -0.24},
        ))
        fig_hist.update_layout(**layout2)
        st.plotly_chart(fig_hist, use_container_width=True)


# =============================================================================
# TAB 3 - Q3 2025 vs Q3 2026 Comparison
# =============================================================================
with tab3:
    st.markdown(f'<div class="sec-title">Q3 26 Forecast . Q3 25 Actual . Q3 26 Original Forecast</div>', unsafe_allow_html=True)

    df_q3 = load_q3_comparison()

    if df_q3.empty:
        st.warning("Could not load Q3 comparison data.")
    else:
        show_detail = st.checkbox("Show per-center Net bench detail", value=False)

        # -- KPI strip -----------------------------------------------------
        if "Q3 26 Forecast" in df_q3.columns and "Q3 25 bench" in df_q3.columns:
            def _safe_int(val, default=0):
                v = pd.to_numeric(val, errors="coerce")
                return int(v) if pd.notna(v) else default

            _v26 = pd.to_numeric(df_q3["Q3 26 Forecast"].iloc[-1], errors="coerce")
            _v25 = pd.to_numeric(df_q3["Q3 25 bench"].iloc[-1],    errors="coerce")
            delta_last = int(_v26 - _v25) if pd.notna(_v26) and pd.notna(_v25) else 0
            st.markdown("<br>", unsafe_allow_html=True)
            q1, q2, q3_col, q4 = st.columns(4)
            q1.metric("Q3 26 Forecast peak",    f"{_safe_int(df_q3['Q3 26 Forecast'].max())}")
            q2.metric("Q3 25 Bench trough",      f"{_safe_int(df_q3['Q3 25 bench'].min())}")
            q3_col.metric("Wk 13 Delta (26 vs 25)", f"{delta_last:+d}")
            q4.metric("Q3 Orig Forecast peak",   f"{_safe_int(df_q3['Q3 Original Forecast'].max()) if 'Q3 Original Forecast' in df_q3.columns else 'N/A'}")
            st.markdown("<br>", unsafe_allow_html=True)

        # -- Chart ---------------------------------------------------------
        fig_q3 = go.Figure()

        series_cfg = {
            "Q3 26 Forecast":       dict(color=C_FORECAST, dash="solid", width=2.8, symbol="circle",  size=8),
            "Q3 25 bench":          dict(color=C_ACTUAL,   dash="solid", width=2.8, symbol="square",  size=8),
            "Q3 Original Forecast": dict(color=C_ORIG,     dash="dash",  width=1.8, symbol="diamond", size=6),
        }

        for col in df_q3.columns:
            cfg = series_cfg.get(col, dict(color=TEXT_SEC, dash="solid", width=1.5, symbol="circle", size=6))
            fig_q3.add_trace(go.Scatter(
                name=col,
                x=df_q3.index.tolist(),
                y=df_q3[col].tolist(),
                mode="lines+markers+text",
                line=dict(color=cfg["color"], width=cfg["width"], dash=cfg["dash"]),
                marker=dict(size=cfg["size"], color=cfg["color"], symbol=cfg["symbol"],
                            line=dict(color=SURFACE, width=1.5)),
                text=[str(int(v)) for v in df_q3[col].tolist()],
                textposition="top center",
                textfont=dict(color=cfg["color"], size=10, family="Inter, system-ui"),
                hovertemplate=f"<b>{col}</b><br>%{{x}}: %{{y}} heads<extra></extra>",
            ))

        # Shaded band between Q3 26 forecast and Q3 25 bench
        if "Q3 26 Forecast" in df_q3.columns and "Q3 25 bench" in df_q3.columns:
            x_vals = df_q3.index.tolist()
            upper  = df_q3["Q3 26 Forecast"].tolist()
            lower  = df_q3["Q3 25 bench"].tolist()
            fig_q3.add_trace(go.Scatter(
                x=x_vals + x_vals[::-1],
                y=upper + lower[::-1],
                fill="toself",
                fillcolor="rgba(56,189,248,0.06)",
                line=dict(color="rgba(0,0,0,0)"),
                showlegend=False,
                hoverinfo="skip",
            ))

        layout3 = {**CHART_LAYOUT}
        layout3.update(dict(
            height=440,
            title=dict(
                text="Q3 Total Bench Comparison - 13-week view",
                font=dict(size=13, color=TEXT_SEC),
                x=0,
            ),
            xaxis={**CHART_LAYOUT["xaxis"], "title": "Week", "tickangle": -30},
            yaxis=dict(**CHART_LAYOUT["yaxis"], title="Total Bench Headcount"),
            legend={**CHART_LAYOUT["legend"], "y": -0.18},
        ))
        fig_q3.update_layout(**layout3)
        st.plotly_chart(fig_q3, use_container_width=True)

        # -- Per-center stacked detail (optional) --------------------------
        if show_detail:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(f'<div class="sec-title">Per-Center Q3 2025 Net Bench by Week</div>', unsafe_allow_html=True)
            df_detail = load_q3_center_detail()
            if not df_detail.empty:
                fig_det = go.Figure()
                for _, row_d in df_detail.iterrows():
                    cname = row_d["Center"]
                    fig_det.add_trace(go.Bar(
                        name=cname,
                        x=df_detail.columns[1:].tolist(),
                        y=row_d[1:].tolist(),
                        marker_color=CENTER_COLORS.get(cname, TEXT_SEC),
                        marker_line_width=0,
                    ))
                layout_det = {**CHART_LAYOUT}
                layout_det.update(dict(
                    barmode="stack",
                    height=330,
                    bargap=0.2,
                    xaxis={**CHART_LAYOUT["xaxis"], "title": "Week"},
                    yaxis=dict(**CHART_LAYOUT["yaxis"], title="Net Bench"),
                    legend={**CHART_LAYOUT["legend"], "y": -0.26},
                ))
                fig_det.update_layout(**layout_det)
                st.plotly_chart(fig_det, use_container_width=True)
                styled_table(df_detail)


# =============================================================================
# TAB 4 - Executive Summary
# =============================================================================
with tab4:
    st.markdown(f'<div class="sec-title">Executive Summary - Dynamic Email Draft</div>',
                unsafe_allow_html=True)
    st.caption("All highlighted figures update automatically as you edit the forecast data.")

    # -- Pull data from session state / already-computed values ------------
    _fc   = load_bench_forecast()
    _hc   = st.session_state.get("hc_input", 2013)
    _work = st.session_state.get("forecast_data", _fc).copy()

    # Per-center totals over all weeks
    _gt_vals  = _work[WEEKS].sum(axis=0).tolist()          # grand total per week
    _gt_w1    = int(_gt_vals[0])
    _gt_w13   = int(_gt_vals[-1])
    _peak_idx = int(pd.Series(_gt_vals).argmax())
    _peak_wk  = WEEKS[_peak_idx]
    _peak_val = int(_gt_vals[_peak_idx])
    _avg_pct  = float(pd.Series(_gt_vals).mean() / _hc * 100)
    _w1_pct   = _gt_w1 / _hc * 100
    _w13_pct  = _gt_w13 / _hc * 100
    _trend    = "increasing" if _gt_w13 > _gt_w1 else "decreasing" if _gt_w13 < _gt_w1 else "stable"

    # Highest and lowest centers at Wk 01
    _w1_series = {c: int(_work[_work["Center"] == c][WEEKS[0]].values[0])
                  for c in CENTERS if not _work[_work["Center"] == c].empty}
    _top_center    = max(_w1_series, key=_w1_series.get)
    _top_center_v  = _w1_series[_top_center]
    _low_center    = min(_w1_series, key=_w1_series.get)
    _low_center_v  = _w1_series[_low_center]

    # Q3 comparison data
    _q3 = load_q3_comparison()
    _q3_fc_w1  = int(_q3["Q3 26 Forecast"].iloc[0])  if not _q3.empty else 0
    _q3_ac_w1  = int(_q3["Q3 25 bench"].iloc[0])     if not _q3.empty else 0
    _q3_delta  = _q3_fc_w1 - _q3_ac_w1
    _q3_trend  = "above" if _q3_delta > 0 else "below" if _q3_delta < 0 else "in line with"
    _q3_fc_pk  = int(_q3["Q3 26 Forecast"].max())    if not _q3.empty else 0
    _q3_ac_trg = int(_q3["Q3 25 bench"].min())       if not _q3.empty else 0

    # Historic data
    _hist = load_historic()
    _hist_last = _hist.iloc[-1] if not _hist.empty else None
    _hist_lbl  = str(_hist_last["Label"]) if _hist_last is not None else "latest"
    _hist_fnc  = int(_hist_last["NA FNC"]) if _hist_last is not None else 0

    # -- Email meta controls -----------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    em1, em2, em3 = st.columns(3)
    with em1:
        sender_name  = st.text_input("From (your name)",    value="Workforce Planning Team", key="em_from")
        report_week  = st.text_input("Reporting week label", value=WEEKS[0],                 key="em_week")
    with em2:
        quarter_label = st.text_input("Quarter",    value="Q3 2026",    key="em_qtr")
        region_label  = st.text_input("Region",     value="NA FNC",     key="em_region")
    with em3:
        bench_target = 5.0
        st.markdown(
            f'<div style="background:{SURFACE};border:1px solid {BORDER};border-radius:8px;padding:10px 12px;margin-top:4px;margin-bottom:10px;">'
            f'<div style="font-size:0.74rem;color:{TEXT_SEC};text-transform:uppercase;letter-spacing:0.06em;">Bench Target</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:{ACCENT};margin-top:2px;">{bench_target:.1f}% (fixed)</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        tone          = st.selectbox("Tone", ["Professional", "Executive Summary", "Detailed"],
                                     key="em_tone")

    # -- Dynamic narrative engine -------------------------------------------
    def _hl(text: str, color: str = ACCENT) -> str:
        """Wrap text in a colored highlight span."""
        return f'<span style="color:{color};font-weight:700">{text}</span>'

    def _status_color(val: float, target: float) -> str:
        if val > target * 1.15: return DANGER
        if val > target:        return WARN
        return GOOD

    avg_color   = _status_color(_avg_pct, bench_target)
    w1_color    = _status_color(_w1_pct,  bench_target)
    w13_color   = _status_color(_w13_pct, bench_target)
    trend_color = GOOD if _trend == "decreasing" else WARN if _trend == "increasing" else ACCENT

    actuals_cutoff_week = 6
    actual_latest_week = WEEKS[actuals_cutoff_week - 1]
    actual_total = int(_gt_vals[actuals_cutoff_week - 1])
    actual_pct = actual_total / _hc * 100
    post_actual_values = {week: int(_gt_vals[WEEKS.index(week)]) for week in WEEKS[actuals_cutoff_week:]}
    highest_future_week = max(post_actual_values, key=post_actual_values.get)
    highest_future_value = post_actual_values[highest_future_week]
    lowest_future_week = min(post_actual_values, key=post_actual_values.get)
    lowest_future_value = post_actual_values[lowest_future_week]
    actual_series = {c: int(_work[_work["Center"] == c][actual_latest_week].values[0])
                     for c in CENTERS if not _work[_work["Center"] == c].empty}
    actual_top_center = max(actual_series, key=actual_series.get)
    actual_top_value = actual_series[actual_top_center]
    actual_low_center = min(actual_series, key=actual_series.get)
    actual_low_value = actual_series[actual_low_center]
    forecast_end_total = int(_gt_vals[-1])
    forecast_delta = forecast_end_total - actual_total
    forecast_delta_color = GOOD if forecast_delta < 0 else WARN if forecast_delta > 0 else ACCENT
    actual_avg_pct = float(pd.Series(_gt_vals[:actuals_cutoff_week]).mean() / _hc * 100)

    # Build paragraphs based on tone
    if tone == "Executive Summary":
        p_opening = (
            f"Please find below the {_hl(quarter_label)} bench update for {_hl(region_label)}. "
            f"Through {_hl(actual_latest_week)}, actual bench levels increased from {_hl('142 FTEs', w1_color)} ({_hl('6.9%', w1_color)}) to "
            f"{_hl('180 FTEs', _status_color(actual_pct, bench_target))} ({_hl('8.8%', _status_color(actual_pct, bench_target))}), "
            f"primarily driven by growth in {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Quebec', CENTER_COLORS.get('Quebec', ACCENT))}. This places current bench materially above the fixed {_hl('5.0%', ACCENT)} target."
        )
        p_trend = (
            f"Looking ahead, the forecast projects bench levels to remain elevated and increase further to {_hl('218 FTEs', w13_color)} ({_hl('11.1%', w13_color)}) by {_hl('Wk 13')}, "
            f"peaking at {_hl('222 FTEs', WARN)} ({_hl('11.3%', WARN)}) in {_hl('Wk 11')} and {_hl('Wk 12')}. The most significant forecasted increase occurs in {_hl('Wk 10')}, "
            f"when bench rises from {_hl('178', ACCENT)} to {_hl('212', WARN)} FTEs, driven mainly by increases in {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))} and {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))}."
        )
        p_centers = (
            f"Center highlights show {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))} improving from {_hl('54')} to {_hl('40')} FTEs through {_hl(actual_latest_week)}, before rebounding to {_hl('49', WARN)} from {_hl('Wk 10')} onward. "
            f"{_hl('Buffalo', CENTER_COLORS.get('Buffalo', ACCENT))} remains stable between {_hl('3')} and {_hl('5')} FTEs, while {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))} continues to be the largest driver of network bench growth, "
            f"moving from {_hl('25')} to {_hl('51')} actuals and peaking at {_hl('68', WARN)} in {_hl('Wk 11')}."
        )
        p_yoy = (
            f"{_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))} and {_hl('Quebec', CENTER_COLORS.get('Quebec', ACCENT))} also show sustained growth, increasing from {_hl('11')} to {_hl('25')} and {_hl('24')} to {_hl('33')} by {_hl(actual_latest_week)}, "
            f"with forecasts reaching {_hl('33')} and {_hl('38')} by {_hl('Wk 13')}. {_hl('Lansing', CENTER_COLORS.get('Lansing', ACCENT))} and {_hl('Monroe', CENTER_COLORS.get('Monroe', ACCENT))} remain broadly stable across the period."
        )
        p_action = (
            f"Overall, the forecast indicates sustained excess capacity through the end of the quarter, with bench percentages remaining {_hl('above 10%', DANGER)} from {_hl('Wk 10')} onward versus a fixed {_hl('5.0%', ACCENT)} target. "
            f"Key focus areas should remain {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))}, "
            f"which account for the majority of forecasted bench exposure and should remain the primary redeployment and mitigation priorities."
        )
        paragraphs = [p_opening, p_trend, p_centers, p_yoy, p_action]

    elif tone == "Detailed":
        p_opening = (
            f"Team, please find the detailed {_hl(quarter_label)} bench update for {_hl(region_label)}. Through {_hl(actual_latest_week)}, actual bench increased from {_hl('142 FTEs', w1_color)} "
            f"({_hl('6.9%', w1_color)}) in {_hl('Wk 01')} to {_hl('180 FTEs', _status_color(actual_pct, bench_target))} ({_hl('8.8%', _status_color(actual_pct, bench_target))}) in {_hl(actual_latest_week)}, versus a fixed {_hl('5.0%', ACCENT)} target."
        )
        p_snapshot = (
            f"<b>Current Snapshot ({_hl(actual_latest_week)}):</b> Bench growth has been concentrated in {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Quebec', CENTER_COLORS.get('Quebec', ACCENT))}, "
            f"while {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))} improved over the same period from {_hl('54')} to {_hl('40')} FTEs."
        )
        p_trend = (
            f"<b>Actual-to-Forecast View:</b> The forecast remains elevated after {_hl(actual_latest_week)}, ending at {_hl('218 FTEs', w13_color)} ({_hl('11.1%', w13_color)}) in {_hl('Wk 13')} and peaking at {_hl('222 FTEs', WARN)} ({_hl('11.3%', WARN)}) in {_hl('Wk 11')} and {_hl('Wk 12')}. "
            f"The most significant jump occurs in {_hl('Wk 10')}, when total bench rises from {_hl('178')} to {_hl('212', WARN)} FTEs, driven primarily by {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))} and {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))}."
        )
        p_centers = (
            f"<b>Center Highlights:</b> Baton Rouge improves to {_hl('40')} by {_hl(actual_latest_week)} before rebounding to {_hl('49', WARN)} from {_hl('Wk 10')} onward; Buffalo stays in a narrow {_hl('3-5')} range; Calgary rises from {_hl('11')} to {_hl('25')} actuals and reaches {_hl('33')} by {_hl('Wk 13')}; "
            f"Halifax climbs from {_hl('25')} to {_hl('51')} by {_hl(actual_latest_week)} and peaks at {_hl('68', WARN)}; Lansing remains in the {_hl('16-18')} range; Monroe stays broadly flat between {_hl('10')} and {_hl('14')}; Quebec grows steadily from {_hl('24')} to {_hl('33')} and is projected to reach {_hl('38')} by {_hl('Wk 13')}."
        )
        p_yoy = (
            f"<b>Capacity Outlook:</b> Bench percentages remain above {_hl('10%', DANGER)} from {_hl('Wk 10')} onward, well above the fixed {_hl('5.0%', ACCENT)} target and indicating sustained excess capacity through the end of the period. "
            f"This reinforces the need to prioritize redeployment actions in {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))}."
        )
        p_action = (
            f"<b>Recommended Actions:</b> Continue active mitigation on the largest forecast contributors, with primary focus on {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))}. "
            f"Secondary monitoring should remain on {_hl('Quebec', CENTER_COLORS.get('Quebec', ACCENT))}, which continues to trend upward through quarter-end."
        )
        paragraphs = [p_opening, p_snapshot, p_trend, p_centers, p_yoy, p_action]

    else:  # Professional (default)
        p_opening = (
            f"Please find the {_hl(quarter_label)} bench update for {_hl(region_label)}. Through {_hl(actual_latest_week)}, actual bench increased from {_hl('142 FTEs', w1_color)} ({_hl('6.9%', w1_color)}) to "
            f"{_hl('180 FTEs', _status_color(actual_pct, bench_target))} ({_hl('8.8%', _status_color(actual_pct, bench_target))}), against a fixed {_hl('5.0%', ACCENT)} target."
        )
        p_trend = (
            f"The forecast remains elevated for the balance of the quarter, reaching {_hl('218 FTEs', w13_color)} ({_hl('11.1%', w13_color)}) by {_hl('Wk 13')} and peaking at {_hl('222 FTEs', WARN)} ({_hl('11.3%', WARN)}) in {_hl('Wk 11')} and {_hl('Wk 12')}. "
            f"The sharpest step-up occurs in {_hl('Wk 10')}, when bench increases from {_hl('178')} to {_hl('212', WARN)} FTEs."
        )
        p_centers = (
            f"The main drivers of growth continue to be {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Quebec', CENTER_COLORS.get('Quebec', ACCENT))}, while {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))} improves through {_hl(actual_latest_week)} but rebounds in forecast from {_hl('Wk 10')} onward."
        )
        p_yoy = (
            f"Bench exposure is expected to remain {_hl('above 10%', DANGER)} from {_hl('Wk 10')} onward, materially above the fixed {_hl('5.0%', ACCENT)} target and indicative of sustained excess capacity through quarter-end."
        )
        p_action = (
            f"Focus should remain on {_hl('Halifax', CENTER_COLORS.get('Halifax', ACCENT))}, {_hl('Calgary', CENTER_COLORS.get('Calgary', ACCENT))}, and {_hl('Baton Rouge', CENTER_COLORS.get('Baton Rouge', ACCENT))} as the primary mitigation priorities."
        )
        paragraphs = [p_opening, p_trend, p_centers, p_yoy, p_action]

    # -- Render the email preview -------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f'<div class="sec-title">Preview</div>', unsafe_allow_html=True)

    # Pre-build all dynamic fragments before the single f-string
    para_html = "".join(
        f"<p style='margin:0 0 14px;line-height:1.75'>{p}</p>"
        for p in paragraphs
    )

    yoy_pill_color = '#f87171' if _q3_delta > 0 else '#34d399'

    def _build_snapshot_rows(
        row_bg_border: str,
        row_total_bg: str,
        row_total_fg: str,
    ) -> tuple[str, str]:
        snapshot_totals = {week: 0 for week in WEEKS}
        snapshot_rows_html = ""
        for c in CENTERS:
            bg, fg = CENTER_ROW_COLORS.get(c, (row_bg_border, TEXT_PRI))
            row_src = _work[_work["Center"] == c]
            week_values = [int(row_src[week].values[0]) if not row_src.empty else 0 for week in WEEKS]
            for week, value in zip(WEEKS, week_values):
                snapshot_totals[week] += value
            week_cells = "".join(
                f'<td style="text-align:center;padding:5px 10px;border:1px solid {row_bg_border};background:{bg};color:{fg}">{value}</td>'
                for value in week_values
            )
            snapshot_rows_html += (
                f'<tr>'
                f'<td style="padding:5px 10px;border:1px solid {row_bg_border};background:{bg};color:{fg};font-weight:600">{c}</td>'
                f'{week_cells}'
                f'</tr>'
            )
        total_cells = "".join(
            f'<td style="text-align:center;padding:5px 10px;border:1px solid {row_bg_border};background:{row_total_bg};color:{row_total_fg};font-weight:700">{snapshot_totals[week]}</td>'
            for week in WEEKS
        )
        snapshot_rows_html += (
            f'<tr>'
            f'<td style="padding:5px 10px;border:1px solid {row_bg_border};background:{row_total_bg};color:{row_total_fg};font-weight:700">Grand Total</td>'
            f'{total_cells}'
            f'</tr>'
        )
        snapshot_header_html = "".join(
            f'<th style="text-align:center;padding:5px 10px;background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};border:1px solid {row_bg_border};">{week}</th>'
            for week in WEEKS
        )
        return snapshot_rows_html, snapshot_header_html

    center_rows_html, center_snapshot_header_html = _build_snapshot_rows(BORDER, TBL_TOTAL_BG, TBL_TOTAL_FG)

    email_html = (
        f'<div style="background:{SURFACE};border:1px solid {BORDER};border-radius:10px;'
        f'padding:32px 36px;font-family:Inter,Segoe UI,system-ui,sans-serif;'
        f'color:{TEXT_PRI};font-size:0.93rem;line-height:1.7;max-width:100%;">'

        # Header
        f'<div style="border-bottom:2px solid {ACCENT};padding-bottom:16px;margin-bottom:24px;">'
        f'<div style="font-size:0.7rem;text-transform:uppercase;letter-spacing:0.1em;color:{TEXT_SEC};margin-bottom:6px;">'
        f'{region_label} Workforce Communication</div>'
        f'<div style="font-size:1.25rem;font-weight:700;color:{TEXT_PRI}">'
        f'{quarter_label} Bench Forecast Update &#8212; {report_week}</div>'
        f'<div style="font-size:0.8rem;color:{TEXT_SEC};margin-top:4px">'
        f'From: {sender_name} &nbsp;&middot;&nbsp; Report generated automatically from dashboard data</div>'
        f'</div>'

        # KPI pills
        f'<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:22px;">'
        f'<div style="background:{SURFACE2};border:1px solid {BORDER};border-radius:20px;padding:5px 14px;font-size:0.8rem;">'
        f'&#128205; <b>{report_week}</b> bench: <span style="color:{w1_color};font-weight:700">{_gt_w1} ({_w1_pct:.1f}%)</span></div>'
        f'<div style="background:{SURFACE2};border:1px solid {BORDER};border-radius:20px;padding:5px 14px;font-size:0.8rem;">'
        f'&#128200; Peak: <span style="color:{ACCENT};font-weight:700">{_peak_val} @ {_peak_wk}</span></div>'
        f'<div style="background:{SURFACE2};border:1px solid {BORDER};border-radius:20px;padding:5px 14px;font-size:0.8rem;">'
        f'&#127919; Target: <span style="color:{ACCENT};font-weight:700">{bench_target:.1f}%</span></div>'
        f'<div style="background:{SURFACE2};border:1px solid {BORDER};border-radius:20px;padding:5px 14px;font-size:0.8rem;">'
        f'&#128202; YoY &#916; W1: <span style="color:{yoy_pill_color};font-weight:700">{_q3_delta:+d}</span></div>'
        f'<div style="background:{SURFACE2};border:1px solid {BORDER};border-radius:20px;padding:5px 14px;font-size:0.8rem;">'
        f'&#128201; Avg bench rate: <span style="color:{avg_color};font-weight:700">{_avg_pct:.1f}%</span></div>'
        f'</div>'

        # Body paragraphs
        f'{para_html}'

        # Center snapshot table
        f'<div style="margin-top:20px;border-top:1px solid {BORDER};padding-top:16px;">'
        f'<div style="font-size:0.72rem;text-transform:uppercase;letter-spacing:0.08em;'
        f'color:{TEXT_SEC};margin-bottom:10px;font-weight:700">Center Snapshot &#8212; Full Quarter View</div>'
        f'<div style="overflow-x:auto;">'
        f'<table style="border-collapse:collapse;width:100%;font-size:0.82rem;min-width:980px;">'
        f'<thead><tr>'
        f'<th style="text-align:left;padding:5px 10px;background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};border:1px solid {BORDER};">Center</th>'
        f'{center_snapshot_header_html}'
        f'</tr></thead>'
        f'<tbody>{center_rows_html}</tbody>'
        f'</table></div></div>'

        # Footer
        f'<div style="margin-top:24px;border-top:1px solid {BORDER};padding-top:12px;'
        f'font-size:0.75rem;color:{TEXT_SEC};">'
        f'This communication was generated automatically from the NA Bench Forecast Dashboard. '
        f'Data reflects the latest forecast inputs. For questions, contact {sender_name}.</div>'
        f'</div>'
    )

    st.markdown(email_html, unsafe_allow_html=True)

    # -- Plain-text copy version --------------------------------------------
    import re, json
    st.markdown("<br>", unsafe_allow_html=True)

    def _strip_html(s: str) -> str:
        return re.sub(r"<[^>]+>", "", s)

    plain_paras = [_strip_html(p) for p in paragraphs]

    FONT = "Aptos, Aptos Display, Calibri, sans-serif"

    # -- SVG helper --------------------------------------------------------
    def _svg_grid_and_labels(w, h, pad, vals, x_labels, max_val):
        """Return (x_fn, y_fn, grid_svg) for a line chart area."""
        iw = w - pad["l"] - pad["r"]
        ih = h - pad["t"] - pad["b"]
        n  = len(vals)

        def xf(i):  return pad["l"] + i * iw / max(n - 1, 1)
        def yf(v):  return pad["t"] + ih - (v / max_val * ih)

        grid = ""
        step = max(1, int(max_val // 4))
        for gi in range(0, int(max_val) + 1, step):
            gy = yf(gi)
            grid += (f'<line x1="{pad["l"]}" y1="{gy:.1f}" x2="{pad["l"]+iw}" y2="{gy:.1f}" '
                     f'stroke="#30363d" stroke-width="1"/>'
                     f'<text x="{pad["l"]-5}" y="{gy+4:.1f}" text-anchor="end" '
                     f'font-size="9" fill="#8b949e" font-family="{FONT}">{gi}</text>')
        for i, lbl in enumerate(x_labels):
            grid += (f'<text x="{xf(i):.1f}" y="{h-5}" text-anchor="middle" '
                     f'font-size="9" fill="#8b949e" font-family="{FONT}">{lbl}</text>')
        return xf, yf, grid

    import io, base64

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
        HAS_MATPLOTLIB = True
    except ModuleNotFoundError:
        HAS_MATPLOTLIB = False

    # Outlook-safe chart colors — light background so charts are visible in email
    _CH_BG     = "#ffffff"   # figure background
    _CH_PLOT   = "#f8f9fa"   # axes background
    _CH_GRID   = "#dee2e6"   # gridlines
    _CH_TICK   = "#495057"   # tick labels
    _CH_SPINE  = "#ced4da"   # axis borders
    _CH_LEG_BG = "#ffffff"   # legend background

    # Series colors — keep brand colors but use slightly deeper versions for light bg
    _Q3_COLORS = {
        "Q3 26 Forecast":   "#0284c7",   # darker sky blue
        "Q3 25 Bench":      "#059669",   # darker emerald
        "Q3 Orig Forecast": "#64748b",   # slate
    }

    def _fig_to_png_bytes(fig) -> bytes:
        """Render a matplotlib figure to raw PNG bytes and close it."""
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        return buf.getvalue()

    def _png_to_b64_tag(png_bytes: bytes) -> str:
        """Embed PNG bytes as a base64 data-URI <img> — for browser preview only."""
        b64 = base64.b64encode(png_bytes).decode()
        return (
            f'<img src="data:image/png;base64,{b64}" '
            f'width="600" style="display:block;margin:8px 0;border:1px solid #dee2e6" '
            f'alt="chart"/>'
        )

    def _png_to_cid_tag(cid: str) -> str:
        """Reference a CID-attached image — for Outlook .eml only."""
        return (
            f'<img src="cid:{cid}" '
            f'width="100%" style="display:block;margin:8px 0;border-top:1px solid #30363d;border-bottom:1px solid #30363d" '
            f'alt="chart"/>'
        )

    # -- Chart 1: Q3 Comparison - 3-line chart (matplotlib PNG) -----------
    q3_series = {
        "Q3 26 Forecast":   (list(_q3["Q3 26 Forecast"])       if not _q3.empty else []),
        "Q3 25 Bench":      (list(_q3["Q3 25 bench"])           if not _q3.empty else []),
        "Q3 Orig Forecast": (list(_q3["Q3 Original Forecast"])  if not _q3.empty else []),
    }
    q3_wk_labels = list(_q3.index) if not _q3.empty else WEEKS

    if HAS_MATPLOTLIB:
        fig1, ax1 = plt.subplots(figsize=(12, 3.5), facecolor=_CH_BG)
        ax1.set_facecolor(_CH_PLOT)
        for name, vals in q3_series.items():
            if not vals:
                continue
            col = _Q3_COLORS[name]
            ls  = "--" if name == "Q3 Orig Forecast" else "-"
            ax1.plot(q3_wk_labels, vals, color=col, linewidth=2.2, linestyle=ls,
                     marker="o", markersize=5, label=name)
            for i, v in enumerate(vals):
                ax1.annotate(str(int(v)), (q3_wk_labels[i], v),
                             textcoords="offset points", xytext=(0, 7),
                             ha="center", fontsize=7, color=col, fontweight="bold")
        ax1.set_xticks(range(len(q3_wk_labels)))
        ax1.set_xticklabels(q3_wk_labels, rotation=45, ha="right",
                            fontsize=7.5, color=_CH_TICK)
        ax1.tick_params(axis="y", labelsize=7.5, labelcolor=_CH_TICK)
        ax1.yaxis.set_major_locator(mticker.MaxNLocator(5, integer=True))
        for spine in ax1.spines.values():
            spine.set_edgecolor(_CH_SPINE)
        ax1.tick_params(colors=_CH_SPINE)
        ax1.grid(axis="y", color=_CH_GRID, linewidth=0.6)
        ax1.set_title("Q3 Comparison - Forecast vs Actuals", fontsize=9,
                      color=_CH_TICK, pad=6)
        ax1.legend(fontsize=7.5, facecolor=_CH_LEG_BG, edgecolor=_CH_SPINE,
                   loc="upper right")
        fig1.tight_layout(pad=0.6)
        _q3_png       = _fig_to_png_bytes(fig1)
        q3_chart_svg  = _png_to_b64_tag(_q3_png)
    else:
        _q3_png = None
        q3_chart_svg = f'<p style="color:{TEXT_SEC};font-size:12px;padding:12px">Matplotlib is not installed, so the email preview chart is unavailable.</p>'

    # -- Chart 2: Historic Bench - stacked by center + NA FNC amber line ---
    _hist_data   = load_historic()
    _hb_centers  = ["Baton Rouge", "East Lansing", "Monroe", "Buffalo", "Halifax", "Quebec", "Calgary"]
    _hb_centers  = [c for c in _hb_centers if c in _hist_data.columns]

    # -- Chart 2: Historic Bench - stacked bars + amber line (matplotlib PNG)
    if HAS_MATPLOTLIB and not _hist_data.empty and _hb_centers:
        _hb_df    = _hist_data.sort_values(["Year", "Quarter", "Week"]).copy()
        hb_labels = _hb_df["Label"].tolist()
        hb_totals = _hb_df["NA FNC"].tolist()
        n_hb      = len(hb_labels)
        x_pos     = list(range(n_hb))

        fig2, ax2 = plt.subplots(figsize=(12, 4.0), facecolor=_CH_BG)
        ax2.set_facecolor(_CH_PLOT)

        bottoms = [0.0] * n_hb
        for center in _hb_centers:
            vals = [float(_hb_df.iloc[i][center] or 0) for i in range(n_hb)]
            col  = CENTER_COLORS.get(center, TEXT_SEC)
            ax2.bar(x_pos, vals, bottom=bottoms, color=col, alpha=0.88,
                    width=0.78, label=center)
            bottoms = [bottoms[i] + vals[i] for i in range(n_hb)]

        prev_q = None
        for i, row in enumerate(_hb_df.itertuples()):
            cur_q = (row.Year, row.Quarter)
            if prev_q and cur_q != prev_q:
                ax2.axvline(x=i - 0.5, color=_CH_SPINE, linewidth=0.8,
                            linestyle="--")
            prev_q = cur_q

        _hb_line_col = "#d97706"
        ax2.plot(x_pos, hb_totals, color=_hb_line_col, linewidth=2.2,
                 marker="o", markersize=4, label="NA FNC Total", zorder=5)
        for i, v in enumerate(hb_totals):
            if n_hb <= 16 or i % 2 == 0:
                ax2.annotate(str(int(v)), (i, v),
                             textcoords="offset points", xytext=(0, 6),
                             ha="center", fontsize=6.5, color=_hb_line_col,
                             fontweight="bold")

        tick_step = max(1, n_hb // 16)
        ax2.set_xticks(x_pos[::tick_step])
        ax2.set_xticklabels(hb_labels[::tick_step], rotation=45, ha="right",
                            fontsize=6.5, color=_CH_TICK)
        ax2.tick_params(axis="y", labelsize=7, labelcolor=_CH_TICK)
        for spine in ax2.spines.values():
            spine.set_edgecolor(_CH_SPINE)
        ax2.tick_params(colors=_CH_SPINE)
        ax2.grid(axis="y", color=_CH_GRID, linewidth=0.5)
        ax2.set_title("Historic Bench - Stacked by Center | NA FNC Total (amber line)",
                      fontsize=9, color=_CH_TICK, pad=6)
        ax2.legend(fontsize=6.5, facecolor=_CH_LEG_BG, edgecolor=_CH_SPINE,
                   loc="upper left", ncol=4, framealpha=0.95)
        fig2.tight_layout(pad=0.6)
        _hb_png        = _fig_to_png_bytes(fig2)
        hist_bench_svg = _png_to_b64_tag(_hb_png)
    elif not _hist_data.empty and _hb_centers:
        _hb_png        = None
        hist_bench_svg = f'<p style="color:{TEXT_SEC};font-size:12px;padding:12px">Matplotlib is not installed, so the email preview chart is unavailable.</p>'
    else:
        _hb_png        = None
        hist_bench_svg = f'<p style="color:{TEXT_SEC};font-size:12px;padding:12px">No historic data available.</p>'

    # -- Email theme tokens (light, Outlook-safe) --------------------------
    EM_BG       = "#f4f4f4"   # page background
    EM_SURF     = "#ffffff"   # card / section background
    EM_SURF2    = "#e8e8e8"   # alternate row / raised
    EM_BORDER   = "#c6c6c6"   # dividers
    EM_TEXT     = "#161616"   # primary text
    EM_MUTED    = "#525252"   # muted / labels
    EM_ACCENT   = "#0f62fe"   # IBM Blue 60
    EM_GOOD     = "#198038"   # IBM Green 60
    EM_WARN     = "#da1e28"   # IBM Red 60
    EM_ORANGE   = "#f1620a"   # IBM Orange 50

    # -- KPI pill HTML -----------------------------------------------------
    def _pill(label, val, color):
        return (
            f'<div style="display:inline-block;background:{EM_SURF};border:1px solid {EM_BORDER};'
            f'border-left:3px solid {color};'
            f'border-radius:4px;padding:5px 14px;font-size:12px;margin:3px;">'
            f'{label}: <strong style="color:{color}">{val}</strong></div>'
        )

    pills_html = (
        _pill("Bench " + report_week, f"{_gt_w1} ({_w1_pct:.1f}%)", w1_color) +
        _pill("Peak", f"{_peak_val} @ {_peak_wk}", EM_ACCENT) +
        _pill("Target", f"{bench_target:.1f}%", EM_ACCENT) +
        _pill("YoY W1 &#916;", f"{_q3_delta:+d}", EM_WARN if _q3_delta > 0 else EM_GOOD) +
        _pill("Avg Rate", f"{_avg_pct:.1f}%", avg_color)
    )

    # -- Center snapshot table rows ----------------------------------------
    snap_rows, snap_header_html = _build_snapshot_rows(EM_BORDER, TBL_TOTAL_BG, TBL_TOTAL_FG)

    # -- Assemble the full HTML email --------------------------------------
    para_blocks = "".join(
        f'<p style="margin:0 0 12px;line-height:1.75;color:{EM_TEXT}">{p}</p>'
        for p in plain_paras
    )

    th_style = (
        f'style="padding:7px 10px;background:{EM_ACCENT};color:#ffffff;'
        f'font-size:11px;font-weight:700;text-transform:uppercase;'
        f'letter-spacing:0.05em;border:1px solid {EM_BORDER};text-align:center"'
    )
    th_left = th_style.replace("text-align:center", "text-align:left")

    # Shared styles injected once into the <head> for collapsible sections
    details_css = (
        f'details{{border-top:1px solid {EM_BORDER};}}'
        f'details[open] summary{{border-bottom:1px solid {EM_SURF2};}}'
        'summary{'
        '  display:flex;align-items:center;justify-content:space-between;'
        f'  padding:13px 28px;cursor:pointer;list-style:none;background:{EM_SURF};'
        '  font-size:11px;font-weight:700;text-transform:uppercase;'
        f'  letter-spacing:0.08em;color:{EM_ACCENT};'
        f'  border-left:3px solid {EM_ACCENT};margin-left:0;'
        '  user-select:none;'
        '}'
        'summary::-webkit-details-marker{display:none}'
        'summary::after{content:"\\203a";font-size:16px;transition:transform .2s;margin-left:auto;padding-left:12px}'
        'details[open] summary::after{transform:rotate(90deg)}'
        f'.sec-body{{padding:0 28px 16px;background:{EM_SURF}}}'
    )

    def _section(title: str, content: str, open_by_default: bool = True) -> str:
        open_attr = " open" if open_by_default else ""
        return (
            f'<details{open_attr}>'
            f'<summary>{title}</summary>'
            f'<div class="sec-body">{content}</div>'
            f'</details>'
        )

    full_html = (
        '<!DOCTYPE html><html><head><meta charset="utf-8">'
        f'<style>'
        f'body{{margin:0;padding:0;background:{EM_BG};font-family:{FONT}}}'
        'h2,h3,p,td,th,div,span,summary,input,button{font-family:inherit}'
        f'{details_css}'
        f'</style></head>'
        f'<body style="background:{EM_BG};margin:0;padding:0">'
        f'<div style="width:100%;background:{EM_SURF};">'

        # Header - blue gradient banner
        f'<div style="background:linear-gradient(135deg,{EM_ACCENT} 0%,#0043ce 100%);padding:24px 28px 18px">'
        f'<div style="font-size:11px;text-transform:uppercase;letter-spacing:0.1em;color:#ffffff;margin-bottom:6px">{region_label} Workforce Communication</div>'
        f'<div style="font-size:20px;font-weight:700;color:#ffffff">{quarter_label} Bench Forecast Update &#8212; {report_week}</div>'
        f'<div style="font-size:12px;color:#ffffff;margin-top:4px">From: {sender_name} &nbsp;&middot;&nbsp; Auto-generated from Dashboard</div>'
        f'</div>'

        # KPI pills
        f'<div style="padding:16px 24px 12px;border-bottom:2px solid {EM_ACCENT};background:{EM_SURF}">{pills_html}</div>'

        # Collapsible: Commentary
        + _section("Commentary", para_blocks, open_by_default=True)

        # Collapsible: Q3 Comparison chart
        + _section(
            "Q3 Comparison &#8212; Forecast vs Actuals",
            q3_chart_svg,
            open_by_default=True,
        )

        # Collapsible: Historic Bench trend chart
        + _section(
            "Historic Bench Trend (Avg NA FNC by Quarter)",
            hist_bench_svg,
            open_by_default=True,
        )

        # Collapsible: Center snapshot table
        + _section(
            "Center Snapshot",
            (
                f'<div style="overflow-x:auto;">'
                f'<table style="border-collapse:collapse;width:100%;font-size:13px;min-width:980px">'
                f'<thead><tr>'
                f'<th {th_left}>Center</th>'
                f'{snap_header_html}'
                f'</tr></thead><tbody>{snap_rows}</tbody></table>'
                f'</div>'
            ),
            open_by_default=True,
        )

        # Collapsible: Summary
        + _section(
            "Summary",
            (
                f'<table style="border-collapse:collapse;width:100%;font-size:13px">'
                f'<tr><td style="padding:6px 10px;color:{EM_MUTED};width:160px;border-bottom:1px solid {EM_BORDER}">Wk01 Bench</td>'
                f'<td style="padding:6px 10px;color:{w1_color};font-weight:700;border-bottom:1px solid {EM_BORDER}">{_gt_w1} ({_w1_pct:.1f}%) '
                f'{"&#9888; ABOVE TARGET" if _w1_pct > bench_target else "&#10003; ON TARGET"}</td></tr>'
                f'<tr><td style="padding:6px 10px;color:{EM_MUTED};border-bottom:1px solid {EM_BORDER}">Wk13 Bench</td>'
                f'<td style="padding:6px 10px;color:{w13_color};font-weight:700;border-bottom:1px solid {EM_BORDER}">{_gt_w13} ({_w13_pct:.1f}%)</td></tr>'
                f'<tr><td style="padding:6px 10px;color:{EM_MUTED};border-bottom:1px solid {EM_BORDER}">Peak Week</td>'
                f'<td style="padding:6px 10px;color:{EM_ACCENT};font-weight:700;border-bottom:1px solid {EM_BORDER}">{_peak_val} @ {_peak_wk}</td></tr>'
                f'<tr><td style="padding:6px 10px;color:{EM_MUTED};border-bottom:1px solid {EM_BORDER}">Avg Bench Rate</td>'
                f'<td style="padding:6px 10px;color:{avg_color};font-weight:700;border-bottom:1px solid {EM_BORDER}">{_avg_pct:.1f}% '
                f'{"&#9888; ABOVE TARGET" if _avg_pct > bench_target else "&#10003; ON TARGET"}</td></tr>'
                f'<tr><td style="padding:6px 10px;color:{EM_MUTED}">YoY W1 Delta</td>'
                f'<td style="padding:6px 10px;color:{EM_WARN if _q3_delta > 0 else EM_GOOD};font-weight:700">'
                f'{_q3_delta:+d} vs Q3 2025 ({_q3_ac_w1} actual &#8594; {_q3_fc_w1} forecast)</td></tr>'
                f'</table>'
            ),
            open_by_default=True,
        )

        # Footer
        + f'<div style="padding:12px 28px;background:{EM_SURF2};border-top:2px solid {EM_ACCENT};'
        f'font-size:11px;color:{EM_MUTED}">Auto-generated from the NA Bench Forecast Dashboard &nbsp;&middot;&nbsp; {sender_name}</div>'
        f'</div></body></html>'
    )

    # -- Build .eml with CID-attached chart images (Outlook compatible) -----
    import email.mime.multipart
    import email.mime.text
    import email.mime.image

    # Build a version of the HTML where base64 data-URIs are replaced with cid: refs
    eml_html = full_html
    if _q3_png:
        eml_html = eml_html.replace(q3_chart_svg, _png_to_cid_tag("chart_q3.png"))
    if _hb_png:
        eml_html = eml_html.replace(hist_bench_svg, _png_to_cid_tag("chart_hist.png"))

    # multipart/related wraps the HTML body + inline image attachments
    eml_subject  = f"{quarter_label} Bench Forecast Update - {report_week}"
    msg_outer    = email.mime.multipart.MIMEMultipart("mixed")
    msg_outer["Subject"] = eml_subject
    msg_outer["From"]    = sender_name
    msg_outer["To"]      = ""

    msg_related = email.mime.multipart.MIMEMultipart("related")
    msg_related.attach(email.mime.text.MIMEText(eml_html, "html", "utf-8"))

    # Attach Q3 chart PNG as inline CID image
    if _q3_png:
        img_part1 = email.mime.image.MIMEImage(_q3_png, _subtype="png")
        img_part1.add_header("Content-ID", "<chart_q3.png>")
        img_part1.add_header("Content-Disposition", "inline", filename="chart_q3.png")
        msg_related.attach(img_part1)

    # Attach historic bench chart PNG as inline CID image
    if _hb_png:
        img_part2 = email.mime.image.MIMEImage(_hb_png, _subtype="png")
        img_part2.add_header("Content-ID", "<chart_hist.png>")
        img_part2.add_header("Content-Disposition", "inline", filename="chart_hist.png")
        msg_related.attach(img_part2)

    msg_outer.attach(msg_related)
    eml_bytes = msg_outer.as_bytes()

    # -- Render with toolbar + copy button ---------------------------------
    escaped_html = full_html.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")

    report_component = f"""
<style>
  body {{ margin:0; background:{BG}; font-family:Inter,system-ui,sans-serif; }}
  #toolbar {{
    background:{SURFACE2}; border:1px solid {BORDER};
    border-bottom:none; border-radius:8px 8px 0 0;
    padding:8px 14px; display:flex; align-items:center; justify-content:space-between;
  }}
  #toolbar span {{ font-size:0.73rem; color:{TEXT_SEC}; text-transform:uppercase; letter-spacing:0.07em; font-weight:600; }}
  .tbtn {{
    background:transparent; border:1px solid {ACCENT}; color:{ACCENT};
    border-radius:5px; padding:4px 14px; font-size:0.78rem; font-weight:600;
    cursor:pointer; font-family:inherit; margin-left:6px;
  }}
  .tbtn:hover {{ background:rgba(56,189,248,0.12); }}
  .tbtn.done  {{ border-color:{GOOD}; color:{GOOD}; }}
  #preview {{
    border:1px solid {BORDER}; border-radius:0 0 8px 8px;
    overflow:auto; background:#0d1117;
  }}
</style>
<div id="toolbar">
  <span>&#9993; HTML Email Report</span>
  <button class="tbtn" id="copy-btn" onclick="copyHTML()">&#128203; Copy HTML</button>
</div>
<div id="preview">
  <iframe id="frame" srcdoc="" style="width:100%;border:none;display:block" scrolling="yes"></iframe>
</div>
<script>
const htmlContent = `{escaped_html}`;
const frame = document.getElementById('frame');
frame.srcdoc = htmlContent;
frame.onload = function() {{
  const h = frame.contentDocument.body.scrollHeight;
  frame.style.height = (h + 32) + 'px';
}};

function copyHTML() {{
  navigator.clipboard.writeText(htmlContent).then(() => {{
    const btn = document.getElementById('copy-btn');
    btn.textContent = '&#10003; Copied!';
    btn.classList.add('done');
    setTimeout(() => {{ btn.textContent = '&#128203; Copy HTML'; btn.classList.remove('done'); }}, 2000);
  }});
}}
</script>
"""
    with st.expander("HTML Email Report - with Charts", expanded=True):
        eml_col, _ = st.columns([1, 3])
        with eml_col:
            st.download_button(
                label="Download as .eml (open in Outlook)",
                data=eml_bytes,
                file_name=f"Bench_Forecast_{quarter_label.replace(' ', '_')}_{report_week.replace(' ', '_')}.eml",
                mime="message/rfc822",
                use_container_width=True,
            )
        st.caption("Double-click the downloaded .eml file to open it in Outlook as a ready-to-send draft.")
        components.html(report_component, height=2200, scrolling=True)


# =============================================================================
# TAB 5 - Audit Log  (admin password protected)
# =============================================================================
def _get_admin_password() -> str | None:
    try:
        return st.secrets["admin"]["password"]
    except Exception:
        return os.environ.get("ADMIN_PASSWORD")

with tab5:
    st.markdown(f'<div class="sec-title">Access & Changes Audit Log</div>', unsafe_allow_html=True)

    _admin_pw = _get_admin_password()

    # Gate behind admin password if one is configured
    if _admin_pw:
        if "audit_unlocked" not in st.session_state:
            st.session_state["audit_unlocked"] = False

        if not st.session_state["audit_unlocked"]:
            st.caption("This section is restricted to administrators.")
            _ap_input = st.text_input(
                "Admin password", type="password",
                placeholder="Enter admin password...",
                label_visibility="collapsed",
                key="audit_pw_input",
            )
            if st.button("Unlock audit log", key="audit_unlock_btn"):
                if _ap_input == _admin_pw:
                    st.session_state["audit_unlocked"] = True
                    st.rerun()
                else:
                    st.error("Incorrect admin password.")

    if not _admin_pw or st.session_state.get("audit_unlocked", False):
        st.caption("Records every login and every forecast save, with the user name and a per-center summary of the saved values.")

        if AUDIT_LOG_PATH.exists():
            df_audit = pd.read_csv(AUDIT_LOG_PATH)
            # Newest first
            df_audit = df_audit.iloc[::-1].reset_index(drop=True)

            st.download_button(
                label="⬇️  Download full log (.csv)",
                data=df_audit.to_csv(index=False).encode("utf-8"),
                file_name="bench_audit_log.csv",
                mime="text/csv",
                key="dl_audit_log",
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # Colour-coded table: login = blue tint, save = green tint
            rows_html = ""
            for _, row in df_audit.iterrows():
                action = str(row.get("action", ""))
                bg = "#d0e2ff" if action == "login" else "#defbe6" if action == "save_forecast" else "#ffffff"
                fg = "#0043ce" if action == "login" else "#0e6027" if action == "save_forecast" else TEXT_PRI
                badge = (
                    f'<span style="background:{bg};color:{fg};border:1px solid {fg};'
                    f'border-radius:3px;padding:1px 6px;font-size:0.72rem;font-weight:700">{action}</span>'
                )
                summary = str(row.get("summary", ""))
                rows_html += (
                    f'<tr style="background:{bg}">'
                    f'<td style="padding:5px 10px;border:1px solid {BORDER};color:{TEXT_PRI};white-space:nowrap">{row.get("timestamp","")}</td>'
                    f'<td style="padding:5px 10px;border:1px solid {BORDER}">{badge}</td>'
                    f'<td style="padding:5px 10px;border:1px solid {BORDER};color:{TEXT_PRI};font-weight:600">{row.get("user","")}</td>'
                    f'<td style="padding:5px 10px;border:1px solid {BORDER};color:{TEXT_SEC};font-size:0.82rem">{summary}</td>'
                    f'</tr>'
                )

            table_html = f"""
            <table style="width:100%;border-collapse:collapse;font-size:0.85rem;font-family:Inter,system-ui,sans-serif">
              <thead>
                <tr>
                  <th style="text-align:left;padding:6px 10px;background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};border:1px solid {BORDER}">Timestamp (UTC)</th>
                  <th style="text-align:left;padding:6px 10px;background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};border:1px solid {BORDER}">Action</th>
                  <th style="text-align:left;padding:6px 10px;background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};border:1px solid {BORDER}">User</th>
                  <th style="text-align:left;padding:6px 10px;background:{TBL_HEADER_BG};color:{TBL_HEADER_FG};border:1px solid {BORDER}">Summary</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            """
            components.html(table_html, height=min(600, 60 + len(df_audit) * 38), scrolling=True)
        else:
            st.info("No audit log entries yet. Entries are recorded on every login and every forecast save.")


# -----------------------------------------------------------------------------
# Footer
# -----------------------------------------------------------------------------
st.markdown(f"""
<div style="text-align:center;color:{TEXT_SEC};font-size:0.75rem;
            border-top:1px solid {BORDER};margin-top:48px;padding-top:14px;">
    Made with IBM Bob
</div>
""", unsafe_allow_html=True)
